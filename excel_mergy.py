from os import listdir
from numpy import append
from openpyxl import load_workbook, Workbook

files = listdir('D:/python/Excel')

#print(files)

merge_excel = Workbook()
merge_sheet = merge_excel.active

for file in files:
    if file[-4:] != 'xlsx':
        continue
    print(file)

    new_excel = load_workbook('D:/python/Excel/'+ file, read_only = True)
    new_sheet = new_excel.active

    for row in new_sheet.iter_rows():
        row_data = []
        for cell in row:
            #print(cell.value)
            row_data.append(cell.value)
        merge_sheet.append(row_data)
merge_excel.save('D:/python/mergedExcel.xlsx')
merge_excel.close()